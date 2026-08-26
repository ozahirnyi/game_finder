import csv
import hashlib
import json
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from itertools import chain
from pathlib import Path
from typing import Iterable

from fastapi import HTTPException
from openpyxl import load_workbook


MAX_EXPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORTED_GAMES = 500
TITLE_HEADERS = {
    "game",
    "game title",
    "title",
    "product name",
    "content title",
    "game name",
}
GAME_SHEET_MARKERS = {"game", "troph", "purchase", "library", "content", "online", "vr"}
JSON_COLLECTION_KEYS = {"games", "library", "owned games", "owned_games", "items", "data"}


@dataclass(frozen=True)
class PsnExportCandidate:
    title: str
    product_name: str | None = None
    platform: str | None = None
    content_type: str | None = None


def normalize_title(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    title = re.sub(r"\s+", " ", value).strip()
    if not title or len(title) > 255:
        return None
    return title


def _normalized_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _candidate_columns(rows: Iterable[tuple[object, ...]], sheet_name: str) -> tuple[int, list[int]] | None:
    name = sheet_name.lower()
    for row_index, row in enumerate(rows):
        headers = [_normalized_header(value) for value in row]
        columns = [index for index, header in enumerate(headers) if header in TITLE_HEADERS]
        has_game_context = any(marker in name for marker in GAME_SHEET_MARKERS) or any(
            marker in header for header in headers for marker in GAME_SHEET_MARKERS
        )
        if columns and has_game_context:
            return row_index, columns
    return None


def _transaction_detail_columns(
    rows: Iterable[tuple[object, ...]], sheet_name: str
) -> tuple[int, int, int | None, int | None, int | None, int | None] | None:
    if sheet_name.strip().strip('"').casefold() != "transaction detail":
        return None
    for row_index, row in enumerate(rows):
        headers = [_normalized_header(value) for value in row]
        if "game name" in headers:
            return (
                row_index,
                headers.index("game name"),
                headers.index("product name") if "product name" in headers else None,
                headers.index("transaction type") if "transaction type" in headers else None,
                headers.index("content type") if "content type" in headers else None,
                headers.index("platform") if "platform" in headers else None,
            )
    return None


def _validate_content(content: bytes) -> None:
    if not content:
        raise HTTPException(status_code=400, detail="The PSN export file is empty")
    if len(content) > MAX_EXPORT_BYTES:
        raise HTTPException(status_code=413, detail="The PSN export must be 10 MB or smaller")


def _no_game_data_error() -> HTTPException:
    return HTTPException(
        status_code=422,
        detail=(
            "This PSN export was read successfully, but it contains no game activity "
            "or game purchases to import."
        ),
    )


def _finalize_titles(candidates: Iterable[object]) -> list[str]:
    titles: dict[str, str] = {}
    for candidate in candidates:
        title = normalize_title(candidate)
        if title:
            titles.setdefault(title.casefold(), title)
            if len(titles) >= MAX_IMPORTED_GAMES:
                break
    if not titles:
        raise _no_game_data_error()
    return list(titles.values())


def _parse_xlsx_export_candidates(content: bytes) -> list[PsnExportCandidate]:

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Upload the Excel file received from PlayStation (.xlsx)") from exc

    candidates: dict[str, PsnExportCandidate] = {}
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            buffered_rows: list[tuple[object, ...]] = []
            for _ in range(20):
                try:
                    buffered_rows.append(next(rows))
                except StopIteration:
                    break
            transaction_columns = _transaction_detail_columns(buffered_rows, worksheet.title)
            if transaction_columns:
                (
                    header_index,
                    game_name_column,
                    product_name_column,
                    transaction_type_column,
                    content_type_column,
                    platform_column,
                ) = transaction_columns
                for row in chain(buffered_rows[header_index + 1 :], rows):
                    if transaction_type_column is not None:
                        transaction_type = _normalized_header(
                            row[transaction_type_column] if transaction_type_column < len(row) else None
                        )
                        if transaction_type != "product purchase":
                            continue
                    elif content_type_column is not None and _normalized_header(
                        row[content_type_column] if content_type_column < len(row) else None
                    ) != "game":
                        continue
                    title = normalize_title(row[game_name_column] if game_name_column < len(row) else None)
                    if title:
                        product_name = normalize_title(
                            row[product_name_column] if product_name_column is not None and product_name_column < len(row) else None
                        )
                        platform = normalize_title(
                            row[platform_column] if platform_column is not None and platform_column < len(row) else None
                        )
                        content_type = normalize_title(
                            row[content_type_column] if content_type_column is not None and content_type_column < len(row) else None
                        )
                        candidates.setdefault(title.casefold(), PsnExportCandidate(title, product_name, platform, content_type))
                continue

            header = _candidate_columns(buffered_rows, worksheet.title)
            if not header:
                continue
            header_index, columns = header
            for row in chain(buffered_rows[header_index + 1 :], rows):
                for column in columns:
                    title = normalize_title(row[column] if column < len(row) else None)
                    if title:
                        candidates.setdefault(title.casefold(), PsnExportCandidate(title))
                        if len(candidates) >= MAX_IMPORTED_GAMES:
                            return list(candidates.values())
    finally:
        workbook.close()

    if not candidates:
        raise _no_game_data_error()
    return list(candidates.values())


def _parse_xlsx_export(content: bytes) -> list[str]:
    return [candidate.title for candidate in _parse_xlsx_export_candidates(content)]


def _parse_csv_export(content: bytes) -> list[str]:
    try:
        reader = csv.DictReader(StringIO(content.decode("utf-8-sig")), strict=True)
        fieldnames = reader.fieldnames or []
        title_columns = [name for name in fieldnames if _normalized_header(name) in TITLE_HEADERS]
        if not title_columns:
            return []
        return _finalize_titles(row.get(column) for row in reader for column in title_columns)
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail="Upload a UTF-8 CSV PSN export") from exc
    except csv.Error as exc:
        raise HTTPException(status_code=400, detail="Upload a valid CSV PSN export") from exc


def _parse_json_export(content: bytes) -> list[str]:
    try:
        payload = json.loads(content.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=400, detail="Upload a valid JSON PSN export") from exc

    if isinstance(payload, dict):
        payload = next(
            (
                value
                for key, value in payload.items()
                if _normalized_header(key) in JSON_COLLECTION_KEYS and isinstance(value, list)
            ),
            [],
        )
    if not isinstance(payload, list):
        raise HTTPException(status_code=400, detail="Upload a valid JSON PSN export")

    candidates = (
        value
        for item in payload
        if isinstance(item, dict)
        for key, value in item.items()
        if _normalized_header(key) in TITLE_HEADERS
    )
    return _finalize_titles(candidates)


def parse_psn_export_candidates(content: bytes, filename: str | None = None) -> list[PsnExportCandidate]:
    _validate_content(content)
    suffix = Path(filename or "export.xlsx").suffix.casefold()
    if suffix == ".xlsx":
        return _parse_xlsx_export_candidates(content)
    if suffix == ".csv":
        return [PsnExportCandidate(title) for title in _parse_csv_export(content)]
    if suffix == ".json":
        return [PsnExportCandidate(title) for title in _parse_json_export(content)]
    raise HTTPException(status_code=400, detail="Upload a supported PSN export (.xlsx, .csv, or .json)")


def parse_psn_export(content: bytes, filename: str | None = None) -> list[str]:
    return [candidate.title for candidate in parse_psn_export_candidates(content, filename)]


def psn_external_id(title: str) -> str:
    normalized = re.sub(r"\s+", " ", title).strip().casefold()
    digest = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:32]
    return f"psn:{digest}"


def psn_manual_external_id(title: str) -> str:
    normalized = re.sub(r"\s+", " ", title).strip().casefold()
    digest = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:32]
    return f"psn:manual:{digest}"
